import pandas as pd
import json
import os
from typing import List, Dict, Any, Tuple
import numpy as np
import logging
from pathlib import Path
from datetime import datetime
from src.logger_config import get_logger, log_performance
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
from src.therapy_classifier import classify_therapy

def load_keywords_structure() -> Dict[str, List[str]]:
    """Load the keywords structure from the JSON file."""
    with open('data/keywords_structure.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def flatten_keywords_structure(keywords: Dict[str, List[str]]) -> List[str]:
    """Flatten the keywords structure into a single list of columns."""
    return [col for category in keywords.values() for col in category]

def format_yes_no(value: str) -> str:
    """Format fields that should contain YES/NO values."""
    if not value:
        return ""
    
    value_upper = str(value).upper().strip()
    if value_upper in ['YES', 'Y', 'TRUE', '1']:
        return 'YES'
    elif value_upper in ['NO', 'N', 'FALSE', '0']:
        return 'NO'
    else:
        return value_upper

def clean_field_name(field_name: str) -> str:
    """Clean field names by fixing character encoding issues."""
    # Replace common encoding issues
    field_name = field_name.replace('â‰¥', '≥')
    field_name = field_name.replace('â€™', "'")
    field_name = field_name.replace('â€œ', '"')
    field_name = field_name.replace('â€', '"')
    return field_name

def find_matching_field(target_field: str, json_data: dict) -> str:
    """Find matching field in JSON data with flexible matching."""
    # Try exact match first
    if target_field in json_data:
        return json_data[target_field]
    
    # Try case-insensitive match
    target_lower = target_field.lower()
    for key, value in json_data.items():
        if key.lower() == target_lower:
            return value
    
    # Try partial matches for common variations
    field_variations = {
        'publication name': ['publication name', 'publication_name'],
        'line of treatment': ['line of treatment', 'line of therapy'],
        'sponsors': ['sponsors', 'sponsor', 'research sponsor'],
        'trial name': ['trial name', 'trial_name', 'study name'],
    }
    
    target_key = target_lower.replace(' ', '').replace('_', '')
    for pattern, variations in field_variations.items():
        pattern_key = pattern.replace(' ', '').replace('_', '')
        if target_key == pattern_key:
            for var in variations:
                for key, value in json_data.items():
                    if key.lower().replace(' ', '').replace('_', '') == var.replace(' ', '').replace('_', ''):
                        return value
    
    # Return empty string if no match found
    return ""

def format_field_value(field_name: str, value: str) -> Any:
    """Enhanced field formatting with complete coverage of all field types."""
    if not value:
        return np.nan
    
    # Clean the field name first
    field_name = clean_field_name(field_name)
    
    # YES/NO fields
    yes_no_fields = {
        'Relapsed / Refractory',
        'Immune Checkpoint Inhibitor (ICI) Naive',
        'Immune Checkpoint Inhibitor (ICI) failed',
        'Biomarker Inclusion and exclusion criteria'
    }
    
    # Check field type and format accordingly
    if field_name in yes_no_fields:
        return format_yes_no(value)
    else:
        return str(value).strip()

def validate_cancer_type(cancer_type: str) -> str:
    """Validate and standardize cancer type according to requirements."""
    valid_types = [
        'Resected Cutaneous Melanoma',
        'Unresectable Cutaneous Melanoma',
        'Cutaneous melanoma with Brain metastasis',
        'Cutaneous Melanoma with CNS metastasis',
        'Uveal Melanoma',
        'Mucosal Melanoma',
        'Acral Melanoma'
    ]
    
    for valid_type in valid_types:
        if cancer_type.lower() == valid_type.lower():
            return valid_type
    return cancer_type

def validate_stage(stage: str) -> str:
    """Validate and standardize stage according to requirements."""
    valid_stages = [
        'Stage I',
        'Stage I/II',
        'Stage II',
        'Stage II/III',
        'Stage III',
        'Stage III/IV',
        'Stage IV'
    ]
    
    for valid_stage in valid_stages:
        if stage.lower() == valid_stage.lower():
            return valid_stage
    return stage

def validate_line_of_therapy(line: str) -> str:
    """Validate and standardize line of therapy according to requirements."""
    valid_lines = [
        'Neoadjuvant',
        'First Line or Untreated',
        '2nd Line and beyond'
    ]
    
    for valid_line in valid_lines:
        if line.lower() == valid_line.lower():
            return valid_line
    return line

def process_extracted_json(json_data: Dict[str, Any]) -> Tuple[str, List[Dict[str, Any]]]:
    """Process the JSON data from GPT extraction into industry/non-industry lists."""
    sponsor_type = json_data.get('sponsor_type', '')
    treatment_arms = json_data.get('treatment_arms', [])
    
    # Load keywords structure
    keywords = load_keywords_structure()
    columns = flatten_keywords_structure(keywords)
    
    # Clean column names
    columns = [clean_field_name(col) for col in columns]
    
    processed_arms = []
    
    for arm in treatment_arms:
        # Format the data according to requirements
        formatted_arm = {}
        for col in columns:
            # Try both cleaned and original column names
            original_col = col.replace('≥', 'â‰¥')  # Convert back for lookup
            value = arm.get(col, arm.get(original_col, ''))
            formatted_arm[col] = format_field_value(col, value)
        
        # Apply validations
        if 'Cancer Type' in formatted_arm:
            formatted_arm['Cancer Type'] = validate_cancer_type(str(formatted_arm['Cancer Type']))
        if 'Clinical Development Stage' in formatted_arm:
            formatted_arm['Clinical Development Stage'] = validate_stage(str(formatted_arm['Clinical Development Stage']))
        if 'Line of Treatment' in formatted_arm:
            formatted_arm['Line of Treatment'] = validate_line_of_therapy(str(formatted_arm['Line of Treatment']))
        
        processed_arms.append(formatted_arm)
    
    return sponsor_type, processed_arms

def generate_filename_with_date(mode: str, sponsor_type: str = None) -> str:
    """Generate filename with date and mode indicator."""
    today = datetime.now().strftime('%Y-%m-%d')
    
    if mode == 'full_pub':
        return f"{today}_SC_FP"
    elif mode == 'abstract':
        if sponsor_type == "Industry-Sponsored":
            return f"{today}_SC_S"
        elif sponsor_type == "Non Industry-Sponsored":
            return f"{today}_SC_NS"
        else:
            # Default for abstract mode without specific sponsor type
            return f"{today}_SC_abstract"
    else:
        return f"{today}_SC_unknown"

def generate_excel_files(json_data_list: List[Dict[str, Any]], output_dir: str = 'output') -> Tuple[int, int]:
    """Generate two Excel files based on sponsor type from corrected JSON structure."""
    # Load keywords structure
    keywords = load_keywords_structure()
    columns = flatten_keywords_structure(keywords)
    
    # Clean column names
    columns = [clean_field_name(col) for col in columns]
    
    # Initialize data lists
    industry_data = []
    non_industry_data = []
    
    # Process each extraction result
    for json_data in json_data_list:
        sponsor_type, processed_arms = process_extracted_json(json_data)
        
        # Categorize based on sponsor type
        if sponsor_type == "Industry-Sponsored":
            industry_data.extend(processed_arms)
        elif sponsor_type == "Non Industry-Sponsored":
            non_industry_data.extend(processed_arms)
    
    # Create DataFrames with all required columns
    industry_df = pd.DataFrame(industry_data, columns=columns) if industry_data else pd.DataFrame(columns=columns)
    non_industry_df = pd.DataFrame(non_industry_data, columns=columns) if non_industry_data else pd.DataFrame(columns=columns)
    
    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filenames with date and mode
    industry_filename = generate_filename_with_date('abstract', "Industry-Sponsored")
    non_industry_filename = generate_filename_with_date('abstract', "Non Industry-Sponsored")
    
    # Save to Excel files, ensuring they are placed in the correct output directory
    industry_df.to_excel(os.path.join(output_dir, f'{industry_filename}.xlsx'), index=False, engine='openpyxl')
    non_industry_df.to_excel(os.path.join(output_dir, f'{non_industry_filename}.xlsx'), index=False, engine='openpyxl')
    
    return len(industry_data), len(non_industry_data)

def load_keywords_structure_full_pub() -> Dict[str, List[str]]:
    """Load the full publication keywords structure from the JSON file."""
    with open('data/keywords_structure_full_pub.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def flatten_keywords_structure_full_pub(keywords: Dict[str, List[str]]) -> List[str]:
    """Flatten the full_pub keywords structure into a single list of columns."""
    return [col for category in keywords.values() for col in category]

def generate_full_pub_csv_and_excel(json_data_list: List[Dict[str, Any]], output_filename_base: str, output_dir: str = 'output') -> int:
    """
    Processes a list of JSON objects with shared fields and arm-specific fields.
    Each row combines the shared fields with the arm-specific fields for each treatment arm.
    """
    # Load the new keywords structure
    keywords_data = load_keywords_structure_full_pub()
    shared_fields = keywords_data.get('shared_fields', [])
    arm_specific_fields = keywords_data.get('arm_specific_fields', [])
    all_columns = [clean_field_name(col) for col in shared_fields + arm_specific_fields]

    # This will hold the flattened data, with one row per treatment arm.
    flattened_rows = []

    for publication_data in json_data_list:
        # Extract shared fields once (these are the same for all arms)
        shared_data = {
            clean_field_name(field): find_matching_field(field, publication_data)
            for field in shared_fields
        }

        # Process each treatment arm in the publication
        treatment_arms = publication_data.get('treatment_arms', [])
        if not treatment_arms:
            # If there are no arms, still add the shared data as a row
            flattened_rows.append(shared_data)
            continue

        for arm in treatment_arms:
            # Start with the shared data (same for all arms)
            row = shared_data.copy()
            
            # Add the arm-specific data
            for field in arm_specific_fields:
                clean_name = clean_field_name(field)
                value = find_matching_field(field, arm)
                row[clean_name] = format_field_value(clean_name, value)
            
            # Add programmatically classified therapy type
            if 'Generic name' in row:
                row['Type of therapy'] = classify_therapy(row.get('Generic name', ''))
            
            flattened_rows.append(row)

    df = pd.DataFrame(flattened_rows, columns=all_columns)
    os.makedirs(output_dir, exist_ok=True)
    
    csv_path = os.path.join(output_dir, f'{output_filename_base}.csv')
    excel_path = os.path.join(output_dir, f'{output_filename_base}.xlsx')
    
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    # Write Excel without formatting (simple CSV-like format)
    df.to_excel(excel_path, index=False, engine='openpyxl')
    
    return len(df)

def add_qc_to_excel(excel_path: str, qc_results: List[Dict[str, Any]], qc_keywords: List[str]):
    """
    Add QC_Status column and color coding to Excel file. Also add a summary sheet.
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    # Add QC_Status column if not present
    headers = [cell.value for cell in ws[1]]
    if 'QC_Status' not in headers:
        ws.insert_cols(ws.max_column + 1)
        ws.cell(row=1, column=ws.max_column).value = 'QC_Status'
    qc_col_idx = ws.max_column
    # Color mapping
    color_map = {'Green': '90EE90', 'Orange': 'FFD580', 'Red': 'FF7F7F'}
    # Add QC status and color
    for i, qc in enumerate(qc_results, start=2):
        status = qc.get('QC_Status', '')
        ws.cell(row=i, column=qc_col_idx).value = status
        fill = PatternFill(start_color=color_map.get(status, 'FFFFFF'), end_color=color_map.get(status, 'FFFFFF'), fill_type='solid')
        ws.cell(row=i, column=qc_col_idx).fill = fill
    # Add summary sheet
    if 'QC_Summary' not in wb.sheetnames:
        summary = wb.create_sheet('QC_Summary')
        total = len(qc_results)
        green = sum(1 for q in qc_results if q.get('QC_Status') == 'Green')
        orange = sum(1 for q in qc_results if q.get('QC_Status') == 'Orange')
        red = sum(1 for q in qc_results if q.get('QC_Status') == 'Red')
        summary.append(['QC Status', 'Count', 'Percent'])
        for status, count in [('Green', green), ('Orange', orange), ('Red', red)]:
            percent = (count / total * 100) if total else 0
            summary.append([status, count, f"{percent:.1f}%"])
    wb.save(excel_path)

def write_dicts_to_csv(data: list, output_path: str, fieldnames: list = None):
    """
    Write a list of dictionaries to a CSV file. If fieldnames is not provided, use all keys from the first dict.
    """
    if not data:
        return
    if fieldnames is None:
        # Union of all keys in all dicts
        fieldnames = sorted({k for d in data for k in d.keys()})
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in data:
            writer.writerow(row)

class ExcelGenerator:
    def __init__(self):
        self.logger = get_logger(__name__)
        self.output_dir = Path("output")
        self.output_dir.mkdir(exist_ok=True)
        self.logger.info("ExcelGenerator initialized")

    @log_performance
    def generate_reports(self):
        self.logger.info("Starting Excel report generation")
        try:
            industry_file = self.output_dir / "industry_sponsored_trials.xlsx"
            self._generate_industry_report(industry_file)
            self.logger.info(f"Industry sponsored report saved to: {industry_file}")
            non_industry_file = self.output_dir / "non_industry_sponsored_trials.xlsx"
            self._generate_non_industry_report(non_industry_file)
            self.logger.info(f"Non-industry sponsored report saved to: {non_industry_file}")
            self.logger.info("All Excel reports generated successfully")
        except Exception as e:
            self.logger.error(f"Failed to generate Excel reports: {str(e)}", exc_info=True)
            raise

    def _generate_industry_report(self, output_file):
        self.logger.debug("Generating industry sponsored trials report")
        try:
            data = self._fetch_industry_data()
            self.logger.info(f"Retrieved {len(data)} industry sponsored trials")
            self._write_excel_file(data, output_file)
            self.logger.debug(f"Industry report written to {output_file}")
        except Exception as e:
            self.logger.error(f"Error generating industry report: {str(e)}")
            raise

    def _generate_non_industry_report(self, output_file):
        self.logger.debug("Generating non-industry sponsored trials report")
        try:
            data = self._fetch_non_industry_data()
            self.logger.info(f"Retrieved {len(data)} non-industry sponsored trials")
            self._write_excel_file(data, output_file)
            self.logger.debug(f"Non-industry report written to {output_file}")
        except Exception as e:
            self.logger.error(f"Error generating non-industry report: {str(e)}")
            raise